import os
from datetime import datetime

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.utils import get_column_letter

import config

COLUMNS = [
    ("Timestamp",      "scraped_at"),
    ("Source",         "source"),
    ("Rent (฿)",       "rent"),
    ("Move-in Cost",   "move_in_cost"),
    ("Condo Name",     "condo_name"),
    ("Room Type",      "room_type"),
    ("Size (sqm)",     "size_sqm"),
    ("Floor",          "floor"),
    ("Location Tags",  "location_tags"),
    ("Status",         "status"),
    ("Summary",        "summary"),
    ("Post Link",      "post_url"),
]

COL_WIDTHS = {
    "scraped_at":   18,
    "source":       10,
    "rent":         12,
    "move_in_cost": 14,
    "condo_name":   24,
    "room_type":    12,
    "size_sqm":     10,
    "floor":        8,
    "location_tags": 22,
    "status":       20,
    "summary":      60,
    "post_url":     30,
}


def _row_to_values(row) -> list:
    def get(field):
        try:
            return row[field]
        except (IndexError, KeyError):
            return None

    values = []
    for _, field in COLUMNS:
        val = get(field)
        if field in ("rent", "move_in_cost") and val is not None:
            try:
                val = f"{int(val):,}"
            except (ValueError, TypeError):
                pass
        values.append(val)
    return values


def export_excel(rows: list) -> str:
    os.makedirs(config.RESULTS_DIR, exist_ok=True)
    ts = datetime.now().strftime("%Y-%m-%d_%H-%M")
    path = os.path.join(config.RESULTS_DIR, f"rental_{ts}.xlsx")

    rows_data = [_row_to_values(r) for r in rows]
    headers = [h for h, _ in COLUMNS]

    with pd.ExcelWriter(path, engine="openpyxl") as writer:
        pd.DataFrame().to_excel(writer, sheet_name="Listings", index=False)

    wb = load_workbook(path)
    ws = wb["Listings"]

    # Clear placeholder
    for row in ws:
        for cell in row:
            cell.value = None

    # Header
    header_fill = PatternFill("solid", fgColor="2F4F4F")
    header_font = Font(bold=True, color="FFFFFF", size=10)
    ws.append(headers)
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(wrap_text=True, vertical="center")

    # Data rows
    for row_vals in rows_data:
        ws.append(row_vals)

    # Column widths + wrap
    for col_idx, (_, field) in enumerate(COLUMNS, 1):
        col_letter = get_column_letter(col_idx)
        ws.column_dimensions[col_letter].width = COL_WIDTHS.get(field, 14)
        for cell in ws[col_letter]:
            cell.alignment = Alignment(wrap_text=True, vertical="top")

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions

    wb.save(path)
    return path
